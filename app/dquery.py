import csv
import re
import tempfile
from pathlib import Path

import charts


def write_csv_to(rows, path):
    """Write list[dict] to path as UTF-8 BOM CSV. Overwrites if exists."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("(no data)\n", encoding="utf-8-sig")
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_csv(rows):
    """Write to a temp file, return file path string. Used by reports."""
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8-sig", newline=""
    ) as f:
        if rows:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()), extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return f.name


def _sheet_name(sql, index):
    """Derive a short sheet name from SQL and query index."""
    # Extract first table name after FROM/JOIN
    m = re.search(r'\bFROM\s+(\S+)', sql, re.IGNORECASE)
    if not m:
        m = re.search(r'\bJOIN\s+(\S+)', sql, re.IGNORECASE)
    if m:
        # Keep only the last part (after dot) and trim to 20 chars
        tbl = m.group(1).split('.')[-1].strip('`"\' ')[:20]
        name = f"查询{index}_{tbl}"
    else:
        name = f"查询{index}"
    # Excel sheet names max 31 chars, no special chars
    name = re.sub(r'[\\/*?\[\]:]', '_', name)
    return name[:31]


def _coerce_number(value):
    """Convert numeric strings to int/float so Excel charts can reference them."""
    if value is None:
        return ""
    s = str(value).strip()
    if s == "":
        return ""
    num = charts.to_float(s)
    if num is None:
        return value
    if num.is_integer() and "." not in s and "e" not in s.lower():
        return int(num)
    return num


def _estimate_height(text, chars_per_line, max_height):
    """Estimate row height for wrapped text."""
    lines = text.count("\n") + max(1, len(text) // chars_per_line)
    return min(15 * lines, max_height)


def _write_data_sheet(wb, title, rows, sql_text="", conclusion=None):
    """Create one sheet: styled table + native chart + conclusion + SQL block."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title=title)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    if not rows or (len(rows) == 1 and list(rows[0].values()) == ["(no data)"]):
        ws.cell(1, 1, "(no data)")
        return ws

    headers = list(rows[0].keys())

    # Header row with styling
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows; numeric series columns are written as real numbers so
    # native charts can reference them. id-like columns stay strings.
    coerce_cols = set(charts.series_columns(rows))
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            v = row.get(h, "")
            ws.cell(ri, ci, _coerce_number(v) if h in coerce_cols else v)

    # Auto column width (max 40)
    for ci, h in enumerate(headers, 1):
        max_len = max(len(str(h)), max((len(str(r.get(h, ""))) for r in rows), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

    # Native chart anchored to the right of the table (never fatal)
    try:
        ctype = charts.detect_chart_type(rows)
        if ctype:
            anchor = f"{get_column_letter(len(headers) + 2)}2"
            charts.add_native_chart(ws, rows, ctype, anchor)
    except Exception as e:
        print(f"[dquery] chart embed failed: {e}", flush=True)

    # Conclusion and SQL blocks below the table
    cursor = len(rows) + 2  # first blank row after data
    if conclusion:
        label_cell = ws.cell(cursor, 1, "【结论】")
        label_cell.font = Font(bold=True, color="595959")
        body = ws.cell(cursor + 1, 1, conclusion)
        body.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[cursor + 1].height = _estimate_height(conclusion, 50, 200)
        cursor += 2
    if sql_text:
        label_cell = ws.cell(cursor + 1, 1, "【SQL】")
        label_cell.font = Font(bold=True, color="595959")
        sql_cell = ws.cell(cursor + 2, 1, sql_text.strip())
        sql_cell.font = Font(name="Courier New", size=9, color="595959")
        sql_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[cursor + 2].height = min(15 * (sql_text.count('\n') + 1), 200)
    return ws


def combine_to_excel(result_dir, conclusions=None, final_summary=None):
    """
    Combine all query_N.csv files in result_dir into a multi-sheet Excel.
    conclusions: optional list where item i-1 is the text conclusion for query_i.
    final_summary: optional overall summary; when given, a '总结' sheet is prepended.
    Returns the path to the generated xlsx file, or None if no query files found.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    result_dir = Path(result_dir)
    # Collect query_N.csv files in order
    query_files = sorted(
        result_dir.glob("query_*.csv"),
        key=lambda p: int(re.search(r'query_(\d+)', p.stem).group(1))
    )
    if not query_files:
        return None

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    if final_summary:
        sum_ws = wb.create_sheet(title="总结")
        title_cell = sum_ws.cell(1, 1, "最终结论")
        title_cell.font = Font(bold=True, size=12)
        body = sum_ws.cell(2, 1, final_summary)
        body.alignment = Alignment(wrap_text=True, vertical="top")
        sum_ws.column_dimensions["A"].width = 100
        sum_ws.row_dimensions[2].height = _estimate_height(final_summary, 60, 400)

    for i, csv_path in enumerate(query_files, 1):
        # Read CSV
        rows = []
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        # Read SQL label from adjacent .sql file if exists
        sql_file = csv_path.with_suffix('.sql')
        sql_text = sql_file.read_text(encoding="utf-8") if sql_file.exists() else ""
        sheet_name = _sheet_name(sql_text, i)
        conclusion = conclusions[i - 1] if conclusions and i - 1 < len(conclusions) else None
        _write_data_sheet(wb, sheet_name, rows, sql_text=sql_text, conclusion=conclusion)

    if not wb.sheetnames:
        return None

    out_path = result_dir / "result.xlsx"
    wb.save(out_path)
    return str(out_path)


def rows_to_xlsx(rows, summary, title="报表", out_path=None):
    """Build a single-sheet xlsx (table + native chart + conclusion) for fixed reports."""
    import os
    from openpyxl import Workbook

    if out_path is None:
        fd, out_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
    wb = Workbook()
    wb.remove(wb.active)
    safe_title = re.sub(r'[\\/*?\[\]:]', '_', str(title))[:31] or "报表"
    _write_data_sheet(wb, safe_title, rows, conclusion=summary)
    wb.save(out_path)
    return out_path

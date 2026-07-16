import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "app"))

import charts

import pytest


def test_detect_none_for_empty_rows():
    assert charts.detect_chart_type([]) is None


def test_detect_none_without_numeric_column():
    rows = [{"a": "x", "b": "y"}, {"a": "z", "b": "w"}]
    assert charts.detect_chart_type(rows) is None


def test_detect_line_when_first_column_is_date():
    rows = [{"日期": "20260701", "收入": "100"}, {"日期": "20260702", "收入": "200"}]
    assert charts.detect_chart_type(rows) == "line"


def test_detect_line_when_first_column_values_look_like_dates():
    rows = [{"ds": "20260701", "dau": "10"}, {"ds": "20260702", "dau": "20"},
            {"ds": "20260703", "dau": "30"}, {"ds": "20260704", "dau": "40"},
            {"ds": "20260705", "dau": "50"}, {"ds": "20260706", "dau": "60"},
            {"ds": "20260707", "dau": "70"}, {"ds": "20260708", "dau": "80"},
            {"ds": "20260709", "dau": "90"}]
    assert charts.detect_chart_type(rows) == "line"


def test_detect_pie_for_few_categories_single_value_column():
    rows = [{"渠道": f"ch{i}", "收入": str(i * 10 + 10)} for i in range(5)]
    assert charts.detect_chart_type(rows) == "pie"


def test_detect_bar_for_many_categories():
    rows = [{"渠道": f"ch{i}", "收入": str(i)} for i in range(12)]
    assert charts.detect_chart_type(rows) == "bar"


def test_detect_bar_for_multiple_value_columns():
    rows = [{"渠道": "a", "收入": "10", "付费人数": "3"},
            {"渠道": "b", "收入": "20", "付费人数": "5"}]
    assert charts.detect_chart_type(rows) == "bar"


def test_detect_none_for_single_row_single_value():
    # 结果只有一个数据：不出图，文字罗列数字即可
    rows = [{"指标": "昨日充值总额", "金额": "12345"}]
    assert charts.detect_chart_type(rows) is None


def test_detect_none_for_single_row_multi_value():
    # 单行多数值列同样视为"一个数据"（如昨日 DAU/收入/付费人数一行）
    rows = [{"渠道": "全渠道", "收入": "10", "付费人数": "3"}]
    assert charts.detect_chart_type(rows) is None


def test_detect_none_for_single_row_date():
    rows = [{"日期": "20260715", "收入": "100"}]
    assert charts.detect_chart_type(rows) is None


def test_id_columns_excluded_from_series():
    rows = [{"排名": "1", "role_id": "1001", "充值金额": "50"},
            {"排名": "2", "role_id": "1002", "充值金额": "60"}]
    assert charts.series_columns(rows) == ["充值金额"]


def test_varchar_numeric_with_thousands_separator():
    # 游戏 39 场景：数值列是 VARCHAR，可能带千分位
    rows = [{"item": "钻石", "cnt": "1,234"}, {"item": "金币", "cnt": "5,678"}]
    assert charts.series_columns(rows) == ["cnt"]
    assert charts.detect_chart_type(rows) == "pie"


def test_to_float():
    assert charts.to_float("1,234.5") == 1234.5
    assert charts.to_float("42") == 42.0
    assert charts.to_float("abc") is None
    assert charts.to_float(None) is None


def test_slice_for_png_limits():
    rows = [{"c": str(i), "v": str(i)} for i in range(100)]
    assert len(charts._slice_for_png(rows, "pie")) == charts.MAX_PIE_CATEGORIES
    assert len(charts._slice_for_png(rows, "line")) == charts.MAX_LINE_POINTS_PNG
    assert len(charts._slice_for_png(rows, "bar")) == charts.MAX_BAR_ROWS_PNG


_PIE_ROWS = [{"渠道": "甲", "收入": "30"}, {"渠道": "乙", "收入": "70"}]
_LINE_ROWS = [{"日期": f"2026070{i}", "收入": str(i * 100)} for i in range(1, 4)]


def test_render_png_pie(tmp_path):
    if not charts.CHARTS_AVAILABLE:
        pytest.skip("matplotlib not installed")
    out = charts.render_png(_PIE_ROWS, "pie", "充值占比", tmp_path / "pie.png")
    assert out and Path(out).exists() and Path(out).stat().st_size > 0


def test_render_png_line(tmp_path):
    if not charts.CHARTS_AVAILABLE:
        pytest.skip("matplotlib not installed")
    out = charts.render_png(_LINE_ROWS, "line", "收入趋势", tmp_path / "line.png")
    assert out and Path(out).exists() and Path(out).stat().st_size > 0


def test_render_png_bar_grouped(tmp_path):
    if not charts.CHARTS_AVAILABLE:
        pytest.skip("matplotlib not installed")
    rows = [{"渠道": "a", "收入": "10", "付费人数": "3"},
            {"渠道": "b", "收入": "20", "付费人数": "5"}]
    out = charts.render_png(rows, "bar", "分组柱状", tmp_path / "bar.png")
    assert out and Path(out).exists() and Path(out).stat().st_size > 0


def test_render_png_returns_none_when_unavailable(monkeypatch, tmp_path):
    monkeypatch.setattr(charts, "CHARTS_AVAILABLE", False)
    assert charts.render_png(_PIE_ROWS, "pie", "t", tmp_path / "x.png") is None


def test_render_png_returns_none_for_invalid_type(tmp_path):
    assert charts.render_png(_PIE_ROWS, "unknown", "t", tmp_path / "x.png") is None


def test_render_pngs_for_dir(tmp_path):
    (tmp_path / "query_1.csv").write_text("类别,数值\n甲,10\n乙,20\n", encoding="utf-8-sig")
    (tmp_path / "query_1.sql").write_text("SELECT a FROM db.payrecharge", encoding="utf-8")
    (tmp_path / "query_2.csv").write_text("a,b\nx,y\n", encoding="utf-8-sig")  # 无数值列
    paths = charts.render_pngs_for_dir(str(tmp_path))
    if charts.CHARTS_AVAILABLE:
        assert len(paths) == 1
        assert paths[0].endswith("query_1.png")
    else:
        assert paths == []


def test_render_pngs_for_dir_with_gap_in_numbering(tmp_path):
    (tmp_path / "query_1.csv").write_text("类别,数值\n甲,10\n乙,20\n", encoding="utf-8-sig")
    (tmp_path / "query_3.csv").write_text("类别,数值\n丙,30\n丁,40\n", encoding="utf-8-sig")
    paths = charts.render_pngs_for_dir(str(tmp_path))
    if charts.CHARTS_AVAILABLE:
        assert len(paths) == 2
        names = sorted(Path(p).name for p in paths)
        assert names == ["query_1.png", "query_3.png"]
    else:
        assert paths == []


def test_add_native_chart_pie():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["类别", "数值"])
    ws.append(["甲", 10])
    ws.append(["乙", 20])
    rows = [{"类别": "甲", "数值": "10"}, {"类别": "乙", "数值": "20"}]
    ok = charts.add_native_chart(ws, rows, "pie", "D2")
    assert ok is True
    assert len(ws._charts) == 1


def test_add_native_chart_line_multi_series():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["日期", "收入", "付费人数"])
    ws.append(["20260701", 100, 5])
    ws.append(["20260702", 200, 8])
    rows = [{"日期": "20260701", "收入": "100", "付费人数": "5"},
            {"日期": "20260702", "收入": "200", "付费人数": "8"}]
    ok = charts.add_native_chart(ws, rows, "line", "E2")
    assert ok is True
    assert len(ws._charts) == 1


def test_add_native_chart_returns_false_without_series():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append(["x", "y"])
    rows = [{"a": "x", "b": "y"}]
    assert charts.add_native_chart(ws, rows, "bar", "D2") is False
    assert len(ws._charts) == 0


def test_add_native_chart_returns_false_for_invalid_type():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    rows = [{"类别": "甲", "数值": "10"}]
    assert charts.add_native_chart(ws, rows, "unknown", "D2") is False


def test_role_prefixed_metric_columns_are_not_excluded():
    rows = [{"role_level": "1", "role_power": "100", "count": "5"}]
    # role_level is the first (category) column, so it does not become a series.
    # role_power and count are metrics and must not be excluded by the ID regex.
    assert charts.series_columns(rows) == ["role_power", "count"]
    assert not charts._ID_COL_RE.search("role_level")
    assert not charts._ID_COL_RE.search("role_power")
    assert charts._ID_COL_RE.search("role_id")

import sys, csv
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "app"))


def test_write_csv_to_creates_file(tmp_path):
    import dquery
    rows = [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]
    p = tmp_path / "out.csv"
    dquery.write_csv_to(rows, p)
    assert p.exists()


def test_write_csv_to_has_utf8_bom(tmp_path):
    import dquery
    rows = [{"名称": "测试"}]
    p = tmp_path / "out.csv"
    dquery.write_csv_to(rows, p)
    raw = p.read_bytes()
    assert raw[:3] == b'\xef\xbb\xbf', "Missing UTF-8 BOM"


def test_write_csv_to_correct_columns(tmp_path):
    import dquery
    rows = [{"a": "1", "b": "2"}]
    p = tmp_path / "out.csv"
    dquery.write_csv_to(rows, p)
    with open(p, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        data = list(reader)
    assert data[0]["a"] == "1"
    assert data[0]["b"] == "2"


def test_write_csv_empty_rows(tmp_path):
    import dquery
    p = tmp_path / "out.csv"
    dquery.write_csv_to([], p)
    assert p.exists()


def test_write_csv_returns_temp_path():
    import dquery
    rows = [{"x": "1"}]
    path = dquery.write_csv(rows)
    assert path.endswith(".csv")
    assert Path(path).exists()


def _write_query_files(tmp_path, n=1):
    for i in range(1, n + 1):
        (tmp_path / f"query_{i}.csv").write_text(
            "类别,数值\n甲,10\n乙,20\n", encoding="utf-8-sig"
        )
        (tmp_path / f"query_{i}.sql").write_text(
            f"SELECT x FROM db.table{i}", encoding="utf-8"
        )


def test_combine_to_excel_without_new_params_unchanged(tmp_path):
    import dquery
    _write_query_files(tmp_path)
    out = dquery.combine_to_excel(str(tmp_path))
    assert out and Path(out).exists()
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    labels = [ws.cell(r, 1).value for r in range(1, 10)]
    assert "【结论】" not in labels
    # 无结论时 SQL 块位置保持现状：数据 2 行 → 标签在第 5 行
    assert ws.cell(5, 1).value == "【SQL】"


def test_combine_to_excel_writes_conclusion_below_table(tmp_path):
    import dquery
    _write_query_files(tmp_path)
    out = dquery.combine_to_excel(str(tmp_path), conclusions=["甲占三分之一"])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    assert ws.cell(4, 1).value == "【结论】"
    assert ws.cell(5, 1).value == "甲占三分之一"
    # SQL 块随结论下移
    assert ws.cell(7, 1).value == "【SQL】"


def test_combine_to_excel_embeds_native_chart(tmp_path):
    import dquery
    _write_query_files(tmp_path)
    out = dquery.combine_to_excel(str(tmp_path))
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    assert len(ws._charts) == 1


def test_combine_to_excel_coerces_numeric_cells(tmp_path):
    import dquery
    (tmp_path / "query_1.csv").write_text(
        "渠道,收入,role_id\nA,1234.5,10001\n", encoding="utf-8-sig"
    )
    out = dquery.combine_to_excel(str(tmp_path))
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    assert ws.cell(2, 2).value == 1234.5  # 数值列 → 真数字
    assert ws.cell(2, 3).value == "10001"  # id 列 → 保持字符串


def test_combine_to_excel_prepends_summary_sheet(tmp_path):
    import dquery
    _write_query_files(tmp_path, n=2)
    out = dquery.combine_to_excel(
        str(tmp_path), conclusions=["c1", "c2"], final_summary="总体结论"
    )
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames[0] == "总结"
    assert wb["总结"].cell(2, 1).value == "总体结论"
    assert len(wb.sheetnames) == 3


def test_rows_to_xlsx(tmp_path):
    import dquery
    rows = [{"类别": "甲", "数值": "10"}, {"类别": "乙", "数值": "20"}]
    out = dquery.rows_to_xlsx(rows, "结论文字", title="KPI", out_path=str(tmp_path / "r.xlsx"))
    assert Path(out).exists()
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb.active
    assert ws.title == "KPI"
    assert len(ws._charts) == 1
    assert ws.cell(4, 1).value == "【结论】"
    assert ws.cell(5, 1).value == "结论文字"


def test_rows_to_xlsx_empty_rows(tmp_path):
    import dquery
    out = dquery.rows_to_xlsx([], "无数据", out_path=str(tmp_path / "r.xlsx"))
    assert Path(out).exists()

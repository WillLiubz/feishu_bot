"""分析游戏 312 的 5月/6月新服表现,直接产出多 sheet Excel(不保留 CSV)。

5月新服: 1448311606 S1606-Galene-EST (05-04) / 1448311607 S1607-Atreus-EST (05-08)
         1448311608 S1608-Cymothoe-EU (05-13) / 1448311609 S1609-Halie-HKT (05-23)
6月新服: 1448311610 S1610-Panopea-EST (06-04) / 1448311611 S1611-Eunice-EST (06-08)
         1448311612 S1612-Melite-EU (06-13) / 1448311613 S1613-Eulimene-HKT (06-23)
(1448311614 S1614-PST-Reborn-EST 06-30 为特殊重开服,不纳入常规新服对比)

口径(已实测):
- gamelog_raw: game_id = 312(整数), server_id 为 VARCHAR 必须加引号;
- gameeco_raw: game_id = '312'(字符串), server_id 同样加引号;
- rolereg.role_id 为空 -> 注册按 account 去重;
- pay_money 币种为 USD(schema 文档写人民币元,有误);
- payrecharge.pay_type 恒为 1(源码 game_log.go Log_payRecharge 硬编码);
  普通充值 pay_itemid = '0',直购 pay_itemid = 'activityId:giftId'
  (activityId 见 const.pb.go DIRECT_PURCHASE_ACT_ID,14=新手直购);
- 钻石消耗: roleres res_id IN (2=金钻, 3=绑钻), change_type=2=消耗,
  change_reason 对应 GLOG_SOURCE_S 枚举(gamelog.pb.go, 1100 个值)。

输出: debug/output/312_new_servers/312新服分析_5月6月.xlsx
"""
import io
import json
import sys
import time
from collections import defaultdict
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
sys.path.insert(0, str(Path(__file__).parent))

import config
import dataapi
from glog_source_312 import GLOG_SOURCE_PROTO

config.DATA_API_POLL_MAX_ATTEMPTS = 72  # ~6 分钟/查询

OUT_DIR = Path(__file__).parent / "output" / "312_new_servers"
OUT_DIR.mkdir(parents=True, exist_ok=True)

SERVERS_SQL = "('1448311606','1448311607','1448311608','1448311609','1448311610','1448311611','1448311612','1448311613')"
CHUNKS = [("20260504", "20260531"), ("20260601", "20260630"), ("20260701", "20260719")]

SERVER_META = {
    "1448311606": {"batch": "5月", "name": "S1606-Galene-EST", "open": "2026-05-04", "open_ds": "20260504"},
    "1448311607": {"batch": "5月", "name": "S1607-Atreus-EST", "open": "2026-05-08", "open_ds": "20260508"},
    "1448311608": {"batch": "5月", "name": "S1608-Cymothoe-EU", "open": "2026-05-13", "open_ds": "20260513"},
    "1448311609": {"batch": "5月", "name": "S1609-Halie-HKT", "open": "2026-05-23", "open_ds": "20260523"},
    "1448311610": {"batch": "6月", "name": "S1610-Panopea-EST", "open": "2026-06-04", "open_ds": "20260604"},
    "1448311611": {"batch": "6月", "name": "S1611-Eunice-EST", "open": "2026-06-08", "open_ds": "20260608"},
    "1448311612": {"batch": "6月", "name": "S1612-Melite-EU", "open": "2026-06-13", "open_ds": "20260613"},
    "1448311613": {"batch": "6月", "name": "S1613-Eulimene-HKT", "open": "2026-06-23", "open_ds": "20260623"},
}
ORDER = list(SERVER_META.keys())

# 常见 GLOG_SOURCE_S 中文名(覆盖 proto 英文名);未覆盖的用 proto 名兜底
GLOG_SOURCE_CN = {
    101200: "召唤", 101504: "世界树宝箱购买", 101601: "公会捐献",
    102000: "商店购买", 102400: "购买次数", 103400: "活动",
    104200: "点石成金", 106044: "闪光召唤(限时召唤)",
    107330: "招财猫抽钻石", 108292: "拉霸商店", 108332: "跨服转盘商店",
    108523: "买一送一商店", 108630: "砸金蛋玩法",
    118252: "女神方舟商店", 118304: "女神圣诞商店",
    130801: "月度福利活跃消耗", 131232: "水果机商店",
}


def glog_name(rid):
    return GLOG_SOURCE_CN.get(rid) or GLOG_SOURCE_PROTO.get(rid, "")


# 直购活动 ID -> 名称(const.pb.go DIRECT_PURCHASE_ACT_ID)
DIRECT_ACT_NAME = {
    "6": "商店-自动", "7": "商店-普通", "8": "商店-节日",
    "9": "天使通行证", "10": "天使通行证积分", "11": "主题任务",
    "12": "许愿购买", "13": "新月卡", "14": "新手直购",
    "15": "女神市场", "16": "水果通行证", "17": "大亨商店",
    "18": "刮刮卡", "19": "自选礼包", "20": "代金券商店",
    "21": "王者宝藏", "33": "庆典", "34": "女神新市场",
}


def load_newbie_pack_names():
    """从用户提供的新手直购配置表读取 gift_id -> 中文商品名。"""
    cfg = Path(r"C:\Users\liubz\Downloads\新手直购.xlsx")
    names = {}
    if not cfg.exists():
        return names
    from openpyxl import load_workbook

    wb = load_workbook(cfg, read_only=True)
    ws = wb["static_new_player_buy"]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # 第1行: 说明
    next(rows)  # 第2行: 类型
    header = next(rows)  # 第3行: 字段名
    idx_id = header.index("id")
    idx_name = header.index("pack_name")
    for r in rows:
        if r[idx_id] is None:
            continue
        try:
            names[str(r[idx_id])] = json.loads(r[idx_name]).get("cn", "")
        except (json.JSONDecodeError, TypeError):
            pass
    wb.close()
    return names


NEWBIE_PACK_NAME = load_newbie_pack_names()


def run_chunked(title, sql_tpl, chunks=CHUNKS, max_rows=200000):
    """按月分块执行同一 SQL 模板(含 {ds_start}/{ds_end} 占位),合并结果。单块失败重试一次。"""
    all_rows = []
    for ds1, ds2 in chunks:
        sql = sql_tpl.format(ds_start=ds1, ds_end=ds2)
        for attempt in (1, 2):
            try:
                print(f"[query] {title} ({ds1}~{ds2}) ...", flush=True)
                rows = dataapi.run_sql_rows(sql, max_rows=max_rows)
                all_rows.extend(rows)
                break
            except RuntimeError as e:
                if attempt == 2:
                    raise
                print(f"  [retry] {e}", flush=True)
                time.sleep(5)
    return all_rows


# ---------- 1. 每日注册 ----------
daily_reg = run_chunked(
    "每日注册",
    """
SELECT server_id, ds, COUNT(DISTINCT account) AS new_roles
FROM gamelog_raw.v_presto_log_rolereg
WHERE game_id = 312 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
GROUP BY server_id, ds
""".replace("{servers}", SERVERS_SQL),
)
daily_reg.sort(key=lambda r: (str(r["server_id"]), r["ds"]))

# ---------- 2. 付费: 按玩家×类型聚合(pay_type 恒为1, 用 pay_itemid 区分直购) ----------
pay_role = run_chunked(
    "付费(按玩家×类型)",
    """
SELECT server_id, role_id, account,
       CASE WHEN strpos(pay_itemid, ':') > 0 THEN '2' ELSE '1' END AS pay_kind,
       COUNT(*) AS times,
       SUM(COALESCE(TRY_CAST(pay_money AS DOUBLE), 0)) AS money,
       SUM(COALESCE(TRY_CAST(pay_diamond AS BIGINT), 0)) AS diamonds
FROM gamelog_raw.v_presto_log_payrecharge
WHERE game_id = 312 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
GROUP BY server_id, role_id, account,
         CASE WHEN strpos(pay_itemid, ':') > 0 THEN '2' ELSE '1' END
""".replace("{servers}", SERVERS_SQL),
)

# ---------- 3. 直购商品: 按活动×商品×玩家聚合(pay_itemid = 'actId:giftId') ----------
item_role = run_chunked(
    "直购商品(按活动×商品×玩家)",
    """
SELECT server_id,
       split_part(pay_itemid, ':', 1) AS act_id,
       split_part(pay_itemid, ':', 2) AS gift_id,
       role_id,
       COUNT(*) AS times,
       SUM(COALESCE(TRY_CAST(pay_money AS DOUBLE), 0)) AS money,
       SUM(COALESCE(TRY_CAST(pay_diamond AS BIGINT), 0)) AS diamonds
FROM gamelog_raw.v_presto_log_payrecharge
WHERE game_id = 312 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
  AND strpos(pay_itemid, ':') > 0
GROUP BY server_id, split_part(pay_itemid, ':', 1), split_part(pay_itemid, ':', 2), role_id
""".replace("{servers}", SERVERS_SQL),
)

# ---------- 4. 每日付费趋势 ----------
daily_pay = run_chunked(
    "每日付费趋势",
    """
SELECT server_id, ds,
       COUNT(DISTINCT role_id) AS payers,
       ROUND(SUM(COALESCE(TRY_CAST(pay_money AS DOUBLE), 0)), 2) AS money
FROM gamelog_raw.v_presto_log_payrecharge
WHERE game_id = 312 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
GROUP BY server_id, ds
""".replace("{servers}", SERVERS_SQL),
)
daily_pay.sort(key=lambda r: (str(r["server_id"]), r["ds"]))

# ---------- 5. 近7天活跃玩家等级 ----------
level_rows = run_chunked(
    "近7天活跃玩家等级",
    """
SELECT server_id, role_id, MAX(TRY_CAST(role_level AS BIGINT)) AS lvl
FROM gamelog_raw.v_presto_log_rolelogin
WHERE game_id = 312 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
  AND role_type = 1
GROUP BY server_id, role_id
""".replace("{servers}", SERVERS_SQL),
    chunks=[("20260713", "20260719")],
)

# ---------- 6. 钻石消耗去向(金钻+绑钻, 按 change_reason×玩家) ----------
consume_role = run_chunked(
    "钻石消耗去向",
    """
SELECT server_id, change_reason, role_id,
       COUNT(*) AS times,
       SUM(COALESCE(TRY_CAST(change_amount AS BIGINT), 0)) AS cost_diamonds
FROM gameeco_raw.v_presto_log_roleres
WHERE game_id = '312' AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
  AND res_id IN ('2', '3') AND change_type = '2'
  AND role_type = 1
GROUP BY server_id, change_reason, role_id
""".replace("{servers}", SERVERS_SQL),
)

# ---------- 7. 活动参与 ----------
promo_rows = run_chunked(
    "活动参与",
    """
SELECT server_id, activity_topic, activity_pay,
       COUNT(*) AS times,
       COUNT(DISTINCT role_id) AS roles
FROM gameeco_raw.v_presto_log_rolepromo
WHERE game_id = '312' AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
  AND role_type = 1
GROUP BY server_id, activity_topic, activity_pay
""".replace("{servers}", SERVERS_SQL),
)

# ================= 加工 =================
def open_plus7(open_ds):
    import datetime

    y, m, d = int(open_ds[:4]), int(open_ds[4:6]), int(open_ds[6:])
    return (datetime.date(y, m, d) + datetime.timedelta(days=6)).strftime("%Y%m%d")


# 注册
reg_total = {s: 0 for s in ORDER}
first7_reg = {s: 0 for s in ORDER}
for r in daily_reg:
    s = str(r["server_id"])
    reg_total[s] += int(r["new_roles"])
    if SERVER_META[s]["open_ds"] <= r["ds"] <= open_plus7(SERVER_META[s]["open_ds"]):
        first7_reg[s] += int(r["new_roles"])

# 付费汇总 + 类型拆分(role_id 可能为空,空则用 account)
pay_stat = {s: {"roles": set(), "times": 0, "money": 0.0, "diamonds": 0} for s in ORDER}
type_stat = defaultdict(lambda: {"roles": set(), "times": 0, "money": 0.0})
for r in pay_role:
    s = str(r["server_id"])
    pid = r["role_id"] or r["account"]
    st = pay_stat[s]
    st["roles"].add(pid)
    st["times"] += int(r["times"])
    st["money"] += float(r["money"] or 0)
    st["diamonds"] += int(r["diamonds"] or 0)
    t = type_stat[(s, str(r["pay_kind"]))]
    t["roles"].add(pid)
    t["times"] += int(r["times"])
    t["money"] += float(r["money"] or 0)

# 首7日付费
first7_pay = {s: 0.0 for s in ORDER}
for r in daily_pay:
    s = str(r["server_id"])
    if SERVER_META[s]["open_ds"] <= r["ds"] <= open_plus7(SERVER_META[s]["open_ds"]):
        first7_pay[s] += float(r["money"] or 0)

# 商品聚合: (server, act, gift) 及 (server, act)
item_stat = defaultdict(lambda: {"roles": set(), "times": 0, "money": 0.0, "diamonds": 0})
act_stat = defaultdict(lambda: {"roles": set(), "times": 0, "money": 0.0})
for r in item_role:
    s, act, gift = str(r["server_id"]), str(r["act_id"]), str(r["gift_id"])
    it = item_stat[(s, act, gift)]
    it["roles"].add(r["role_id"])
    it["times"] += int(r["times"])
    it["money"] += float(r["money"] or 0)
    it["diamonds"] += int(r["diamonds"] or 0)
    a = act_stat[(s, act)]
    a["roles"].add(r["role_id"])
    a["times"] += int(r["times"])
    a["money"] += float(r["money"] or 0)

# 等级分布
lvl_stat = {s: {"n": 0, "max": 0, "b": [0, 0, 0, 0, 0]} for s in ORDER}
for r in level_rows:
    s = str(r["server_id"])
    lvl = int(r["lvl"] or 0)
    st = lvl_stat[s]
    st["n"] += 1
    st["max"] = max(st["max"], lvl)
    bi = 0 if lvl < 20 else 1 if lvl < 40 else 2 if lvl < 60 else 3 if lvl < 80 else 4
    st["b"][bi] += 1

# 消耗聚合: (server, reason)
consume_stat = defaultdict(lambda: {"roles": set(), "times": 0, "diamonds": 0})
for r in consume_role:
    key = (str(r["server_id"]), int(r["change_reason"]))
    cs = consume_stat[key]
    cs["roles"].add(r["role_id"])
    cs["times"] += int(r["times"])
    cs["diamonds"] += int(r["cost_diamonds"] or 0)


# 活动聚合: (server, topic_cn, activity_pay)
def topic_cn(topic):
    try:
        return json.loads(topic).get("cn") or topic
    except (json.JSONDecodeError, TypeError):
        return topic


promo_stat = defaultdict(lambda: {"times": 0, "roles": 0})
for r in promo_rows:
    s = str(r["server_id"])
    key = (s, topic_cn(r["activity_topic"]), str(r["activity_pay"]))
    ps = promo_stat[key]
    ps["times"] += int(r["times"])
    ps["roles"] += int(r["roles"])  # 分块去重人数累加(近似)

# ================= 汇总行 =================
def batch_sum(stat, batch, key=None):
    total = 0
    for s in ORDER:
        if SERVER_META[s]["batch"] == batch:
            v = stat[s]
            total += v if not key else key(v)
    return total


sheets = {}

sheets["新服概况"] = (
    ["批次", "server_id", "服名", "开服时间", "渠道opgame", "开服规律"],
    [[SERVER_META[s]["batch"], s, SERVER_META[s]["name"], SERVER_META[s]["open"], "1448", "每月4服(4/8/13/23日)"] for s in ORDER],
)

sheets["注册汇总"] = (
    ["批次", "服名", "server_id", "累计注册", "首7日注册"],
    [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s, reg_total[s], first7_reg[s]] for s in ORDER],
)

sheets["付费汇总"] = (
    ["批次", "服名", "server_id", "付费人数", "累计注册", "付费率%", "充值总额(美元)", "充值总钻石", "充值笔数", "ARPPU(美元)"],
    [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s,
      len(pay_stat[s]["roles"]), reg_total[s],
      round(len(pay_stat[s]["roles"]) / reg_total[s] * 100, 1) if reg_total[s] else 0,
      round(pay_stat[s]["money"], 2), pay_stat[s]["diamonds"], pay_stat[s]["times"],
      round(pay_stat[s]["money"] / len(pay_stat[s]["roles"]), 2) if pay_stat[s]["roles"] else 0] for s in ORDER],
)

sheets["首7日对齐对比"] = (
    ["批次", "服名", "server_id", "首7日注册", "首7日付费(美元)", "首7日ARPU(美元)"],
    [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s,
      first7_reg[s], round(first7_pay[s], 2),
      round(first7_pay[s] / first7_reg[s], 2) if first7_reg[s] else 0] for s in ORDER],
)

type_name = {"1": "购买钻石(普通充值)", "2": "购买商品(直购)"}
rows = []
for s in ORDER:
    for pt in ("1", "2"):
        t = type_stat.get((s, pt))
        if t:
            rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s,
                         type_name[pt], len(t["roles"]), t["times"], round(t["money"], 2)])
sheets["付费类型拆分"] = (["批次", "服名", "server_id", "付费类型", "付费人数", "笔数", "金额(美元)"], rows)

sheets["等级分布(近7天活跃)"] = (
    ["批次", "服名", "server_id", "近7天活跃人数", "最高等级", "1-19级", "20-39级", "40-59级", "60-79级", "80级以上", "活跃/累计注册%"],
    [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s,
      lvl_stat[s]["n"], lvl_stat[s]["max"], *lvl_stat[s]["b"],
      round(lvl_stat[s]["n"] / reg_total[s] * 100, 1) if reg_total[s] else 0] for s in ORDER],
)

rows = []
for s in ORDER:
    tops = sorted(
        [(rid, cs) for (srv, rid), cs in consume_stat.items() if srv == s],
        key=lambda kv: -kv[1]["diamonds"],
    )[:15]
    for i, (rid, cs) in enumerate(tops, 1):
        rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s, i,
                     rid, glog_name(rid), len(cs["roles"]), cs["times"], cs["diamonds"]])
sheets["钻石消耗去向TOP15"] = (["批次", "服名", "server_id", "排名", "change_reason", "来源(GLOG_SOURCE_S)", "消耗人数", "消耗次数", "消耗钻石(金钻+绑钻)"], rows)

rows = []
for s in ORDER:
    tops = sorted(
        [(act, a) for (srv, act), a in act_stat.items() if srv == s],
        key=lambda kv: -kv[1]["money"],
    )
    for i, (act, a) in enumerate(tops, 1):
        rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s, i,
                     act, DIRECT_ACT_NAME.get(act, ""), len(a["roles"]), a["times"], round(a["money"], 2)])
sheets["直购活动分布(分服)"] = (["批次", "服名", "server_id", "排名", "活动ID", "活动名称", "付费人数", "购买笔数", "金额(美元)"], rows)

rows = []
for s in ORDER:
    tops = sorted(
        [(act, gift, it) for (srv, act, gift), it in item_stat.items() if srv == s],
        key=lambda kv: -kv[2]["money"],
    )[:15]
    for i, (act, gift, it) in enumerate(tops, 1):
        pname = NEWBIE_PACK_NAME.get(gift, "") if act == "14" else ""
        rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s, i,
                     DIRECT_ACT_NAME.get(act, act), gift, pname,
                     len(it["roles"]), it["times"], round(it["money"], 2), it["diamonds"]])
sheets["直购商品TOP15(分服)"] = (["批次", "服名", "server_id", "排名", "直购活动", "商品ID", "商品名", "付费人数", "购买笔数", "金额(美元)", "购得钻石"], rows)

rows = []
for s in ORDER:
    tops = sorted(
        [(gift, it) for (srv, act, gift), it in item_stat.items() if srv == s and act == "14"],
        key=lambda kv: -kv[1]["money"],
    )
    for i, (gift, it) in enumerate(tops, 1):
        rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s, i, gift,
                     NEWBIE_PACK_NAME.get(gift, ""),
                     len(it["roles"]), it["times"], round(it["money"], 2)])
sheets["新手直购明细(分服)"] = (["批次", "服名", "server_id", "排名", "商品ID", "商品名", "付费人数", "购买笔数", "金额(美元)"], rows)

rows = []
for s in ORDER:
    tops = sorted(
        [(topic, pay, ps) for (srv, topic, pay), ps in promo_stat.items() if srv == s],
        key=lambda kv: -kv[2]["roles"],
    )[:15]
    for i, (topic, pay, ps) in enumerate(tops, 1):
        rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s, i, topic,
                     "是" if pay == "1" else "否", ps["roles"], ps["times"]])
sheets["活动参与TOP15"] = (["批次", "服名", "server_id", "排名", "活动(中文名)", "是否充值活动", "参与人数(分块累加)", "参与次数"], rows)

sheets["每日新增注册"] = (
    ["批次", "服名", "server_id", "日期(ds)", "新增注册"],
    [[SERVER_META[str(r["server_id"])]["batch"], SERVER_META[str(r["server_id"])]["name"], r["server_id"],
      r["ds"], r["new_roles"]] for r in daily_reg],
)

sheets["每日付费趋势"] = (
    ["批次", "服名", "server_id", "日期(ds)", "付费人数", "付费金额(美元)"],
    [[SERVER_META[str(r["server_id"])]["batch"], SERVER_META[str(r["server_id"])]["name"], r["server_id"],
      r["ds"], r["payers"], r["money"]] for r in daily_pay],
)

# ================= 写 xlsx =================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")

wb = Workbook()
wb.remove(wb.active)
for sheet_name, (header, rows) in sheets.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.append(header)
    for r in rows:
        ws.append(r)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(ws.columns, 1):
        width = max((sum(2 if ord(ch) > 127 else 1 for ch in str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 40)

xlsx_path = OUT_DIR / "312新服分析_5月6月.xlsx"
wb.save(xlsx_path)
print(f"\n合并完成 -> {xlsx_path}")

# ================= 摘要 =================
print("\n===== 摘要 =====")
for s in ORDER:
    m = SERVER_META[s]
    p = pay_stat[s]
    print(
        f"{m['batch']} {m['name']}: 注册={reg_total[s]} 首7日={first7_reg[s]} | "
        f"付费{len(p['roles'])}人 ${round(p['money'], 2)} | 首7日付费=${round(first7_pay[s], 2)} | "
        f"近7天活跃={lvl_stat[s]['n']} 最高级={lvl_stat[s]['max']}"
    )
for batch in ("5月", "6月"):
    regs = sum(reg_total[s] for s in ORDER if SERVER_META[s]["batch"] == batch)
    money = sum(pay_stat[s]["money"] for s in ORDER if SERVER_META[s]["batch"] == batch)
    payers = len(set().union(*[pay_stat[s]["roles"] for s in ORDER if SERVER_META[s]["batch"] == batch]))
    f7r = sum(first7_reg[s] for s in ORDER if SERVER_META[s]["batch"] == batch)
    f7p = sum(first7_pay[s] for s in ORDER if SERVER_META[s]["batch"] == batch)
    print(f"[{batch}合计] 注册={regs} 首7日={f7r} | 付费{payers}人 ${round(money, 2)} | 首7日付费=${round(f7p, 2)} 首7日ARPU=${round(f7p / f7r, 2) if f7r else 0}")
    dp = sum(a["money"] for (s, act), a in act_stat.items() if SERVER_META[s]["batch"] == batch)
    nb = sum(a["money"] for (s, act), a in act_stat.items() if SERVER_META[s]["batch"] == batch and act == "14")
    nb_roles = set()
    for (s, act), a in act_stat.items():
        if SERVER_META[s]["batch"] == batch and act == "14":
            nb_roles |= a["roles"]
    print(f"[{batch}直购] 直购合计=${round(dp, 2)} | 新手直购 {len(nb_roles)}人 ${round(nb, 2)}")

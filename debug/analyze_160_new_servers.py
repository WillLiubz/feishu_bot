"""分析游戏 160 的 5月/6月新服表现,并导出 CSV + 多 sheet Excel。

5月新服: 2251312168 S2168-Tempest-EST (2026-05-06)
6月新服: 2251312169 S2169-Angel-GMT  (2026-06-06)

分析窗口: 2026-05-06 ~ 2026-07-19
策略: gamelog_raw 只按 ds 分区,全窗口扫描超时,故按月分块查询后在 Python 合并。
输出目录: debug/output/160_new_servers/(只保留合并 xlsx)

口径(2026-07 源码确认): payrecharge.pay_type 恒为 1(vip.go 两处 PayRechargeLog
均传 PAY_RECHARGE_BUY_DIAMOND);直购看 pay_itemid:普通充值='0',
直购='actId:giftId'(源码 fmt 参数类型错误,数仓中为 '%!d(string=3:8001)' 形态,
解析时需剥掉 '%!d(string=' 前缀与 ')' 后缀)。
"""
import csv
import io
import sys
import time
from collections import defaultdict
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent / "app"))

import config
import dataapi

config.DATA_API_POLL_MAX_ATTEMPTS = 72  # ~6 分钟/查询

OUT_DIR = Path(__file__).parent / "output" / "160_new_servers"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# server_id 是 VARCHAR,必须加引号(不加引号会导致全表扫描超时,已实测验证)
SERVERS_SQL = "('2251312168', '2251312169')"
DS_START, DS_END = "20260506", "20260719"
CHUNKS = [("20260506", "20260531"), ("20260601", "20260630"), ("20260701", "20260719")]

SERVER_META = {
    "2251312168": {"batch": "5月", "name": "S2168-Tempest-EST", "open": "2026-05-06", "open_ds": "20260506"},
    "2251312169": {"batch": "6月", "name": "S2169-Angel-GMT", "open": "2026-06-06", "open_ds": "20260606"},
}
ORDER = list(SERVER_META.keys())

sys.path.insert(0, str(Path(__file__).parent))
from glog_source_160 import GLOG_SOURCE_PROTO

# 常见来源的中文名(覆盖 proto 英文名);未覆盖的用 proto 名兜底
GLOG_SOURCE_CN = {
    13: "装备强化", 15: "装备精炼", 18: "普通抽奖", 19: "超级抽奖", 26: "商店",
    29: "通天塔", 40: "商店刷新", 47: "商城", 48: "重置关卡", 52: "竞技场清除CD",
    53: "竞技场购买次数", 55: "重置", 58: "购买体力", 63: "竞技场换敌", 65: "竞技场重置",
    84: "宝物强化", 85: "活动", 86: "购买挂机精力", 88: "七日商店", 90: "英雄练习购买次数",
    96: "通天塔商店", 97: "神秘商店", 98: "竞技场商店", 99: "挂机商店", 100: "家族商店",
    108: "宝物一键强化", 112: "VIP奖励", 120: "购买家族BOSS次数", 125: "首充",
    136: "购买次数", 138: "超值购购买", 139: "超值购充值", 141: "超级抽奖",
    153: "开基金购买", 175: "商城首页", 246: "团购购买", 251: "购买世界BOSS次数",
    256: "转盘单抽", 257: "转盘十连", 280: "幸运树", 300: "基金", 301: "基金", 302: "基金",
    306: "扫雷", 352: "端午制作", 390: "爬塔跳层", 416: "跨服转盘十连", 430: "卡牌抽卡",
    1139: "DW加奖(DW_ADD_AWARD)", 10030: "双倍购买(BUY_DOUBLE_BUY)",
}


# 直购活动 ID -> 名称(game.pb.go DIRECT_TYPE 枚举)
DIRECT_ACT_NAME = {
    "1": "活动直购", "2": "商店直购", "3": "新月卡",
    "4": "周年新通行证", "5": "周年新加油站", "6": "基金新通行证",
    "7": "刮刮卡", "8": "代金券", "9": "自定义礼包",
    "10": "王之财宝", "11": "渔场通行证",
}


def glog_name(bid):
    return GLOG_SOURCE_CN.get(bid) or GLOG_SOURCE_PROTO.get(bid, "")


def run_chunked(title, sql_tpl, chunks=CHUNKS, max_rows=100000):
    """按月分块执行同一 SQL 模板(含 {ds_start}/{ds_end} 占位),合并结果。单块失败重试一次。"""
    all_rows = []
    for ds1, ds2 in chunks:
        sql = sql_tpl.format(ds_start=ds1, ds_end=ds2)
        for attempt in (1, 2):
            try:
                print(f"[query] {title} ({ds1}~{ds2}) ...", flush=True)
                rows = dataapi.run_sql_rows(sql, max_rows=max_rows)
                all_rows.extend(rows)
                break
            except RuntimeError as e:
                if attempt == 2:
                    raise
                print(f"  [retry] {e}", flush=True)
                time.sleep(5)
    return all_rows


def write_csv(filename, header, rows):
    path = OUT_DIR / filename
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"[OK] {path.name}  ({len(rows)} 行)", flush=True)


# ---------- 1. 每日注册 ----------
daily_reg = run_chunked(
    "每日注册",
    """
SELECT server_id, ds, COUNT(DISTINCT account) AS new_roles
FROM gamelog_raw.v_presto_log_rolereg
WHERE game_id = 160 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
GROUP BY server_id, ds
""".replace("{servers}", SERVERS_SQL),
)
daily_reg.sort(key=lambda r: (str(r["server_id"]), r["ds"]))

# ---------- 2. 付费: 按玩家×类型聚合(pay_type 恒为1, 用 pay_itemid 区分直购) ----------
pay_role = run_chunked(
    "付费(按玩家×类型)",
    """
SELECT server_id, role_id,
       CASE WHEN strpos(pay_itemid, ':') > 0 THEN '2' ELSE '1' END AS pay_kind,
       COUNT(*) AS times,
       SUM(COALESCE(TRY_CAST(pay_money AS DOUBLE), 0)) AS money,
       SUM(COALESCE(TRY_CAST(pay_diamond AS BIGINT), 0)) AS diamonds
FROM gamelog_raw.v_presto_log_payrecharge
WHERE game_id = 160 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
GROUP BY server_id, role_id,
         CASE WHEN strpos(pay_itemid, ':') > 0 THEN '2' ELSE '1' END
""".replace("{servers}", SERVERS_SQL),
)

# ---------- 3. 直购商品: 按活动×商品×玩家聚合 ----------
# 160 直购 pay_itemid = 'actId:giftId', 但源码 fmt 参数类型错误,
# 数仓中实际为 '%!d(string=3:8001)' 形态, 需剥掉 '%!d(string=' 前缀和 ')' 后缀
item_role = run_chunked(
    "直购商品(按活动×商品×玩家)",
    """
SELECT server_id,
       replace(split_part(pay_itemid, ':', 1), '%!d(string=', '') AS act_id,
       replace(split_part(pay_itemid, ':', 2), ')', '') AS gift_id,
       role_id,
       COUNT(*) AS times,
       SUM(COALESCE(TRY_CAST(pay_money AS DOUBLE), 0)) AS money,
       SUM(COALESCE(TRY_CAST(pay_diamond AS BIGINT), 0)) AS diamonds
FROM gamelog_raw.v_presto_log_payrecharge
WHERE game_id = 160 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
  AND strpos(pay_itemid, ':') > 0
GROUP BY server_id,
         replace(split_part(pay_itemid, ':', 1), '%!d(string=', ''),
         replace(split_part(pay_itemid, ':', 2), ')', ''),
         role_id
""".replace("{servers}", SERVERS_SQL),
)

# ---------- 4. 每日付费趋势 ----------
daily_pay = run_chunked(
    "每日付费趋势",
    """
SELECT server_id, ds,
       COUNT(DISTINCT role_id) AS payers,
       ROUND(SUM(COALESCE(TRY_CAST(pay_money AS DOUBLE), 0)), 2) AS money
FROM gamelog_raw.v_presto_log_payrecharge
WHERE game_id = 160 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
GROUP BY server_id, ds
""".replace("{servers}", SERVERS_SQL),
)
daily_pay.sort(key=lambda r: (str(r["server_id"]), r["ds"]))

# ---------- 5. 近7天活跃玩家等级 ----------
level_rows = run_chunked(
    "近7天活跃玩家等级",
    """
SELECT server_id, role_id, MAX(TRY_CAST(role_level AS BIGINT)) AS lvl
FROM gamelog_raw.v_presto_log_rolelogin
WHERE game_id = 160 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
GROUP BY server_id, role_id
""".replace("{servers}", SERVERS_SQL),
    chunks=[("20260713", "20260719")],
)

# ---------- 6. 钻石消耗去向(按 b_id×玩家) ----------
consume_role = run_chunked(
    "钻石消耗去向",
    """
SELECT server_id, b_id, role_id,
       COUNT(*) AS times,
       SUM(COALESCE(TRY_CAST(consume_diamond AS BIGINT), 0) + COALESCE(TRY_CAST(consume_blackdiamond AS BIGINT), 0)) AS cost_diamonds
FROM gamelog_raw.v_presto_log_payconsume
WHERE game_id = 160 AND server_id IN {servers}
  AND ds BETWEEN '{ds_start}' AND '{ds_end}'
GROUP BY server_id, b_id, role_id
""".replace("{servers}", SERVERS_SQL),
)

# ================= 加工 =================
def open_plus7(open_ds):
    import datetime

    y, m, d = int(open_ds[:4]), int(open_ds[4:6]), int(open_ds[6:])
    return (datetime.date(y, m, d) + datetime.timedelta(days=6)).strftime("%Y%m%d")


# 注册
reg_total = {s: 0 for s in ORDER}
first7_reg = {s: 0 for s in ORDER}
for r in daily_reg:
    s = str(r["server_id"])
    reg_total[s] += int(r["new_roles"])
    if SERVER_META[s]["open_ds"] <= r["ds"] <= open_plus7(SERVER_META[s]["open_ds"]):
        first7_reg[s] += int(r["new_roles"])

# 付费汇总 + 类型拆分
pay_stat = {s: {"roles": set(), "times": 0, "money": 0.0, "diamonds": 0} for s in ORDER}
type_stat = defaultdict(lambda: {"roles": set(), "times": 0, "money": 0.0})
for r in pay_role:
    s = str(r["server_id"])
    st = pay_stat[s]
    st["roles"].add(r["role_id"])
    st["times"] += int(r["times"])
    st["money"] += float(r["money"] or 0)
    st["diamonds"] += int(r["diamonds"] or 0)
    t = type_stat[(s, str(r["pay_kind"]))]
    t["roles"].add(r["role_id"])
    t["times"] += int(r["times"])
    t["money"] += float(r["money"] or 0)

# 首7日付费
first7_pay = {s: 0.0 for s in ORDER}
for r in daily_pay:
    s = str(r["server_id"])
    if SERVER_META[s]["open_ds"] <= r["ds"] <= open_plus7(SERVER_META[s]["open_ds"]):
        first7_pay[s] += float(r["money"] or 0)

# 商品聚合: (server, act, gift) 及 (server, act)
item_stat = defaultdict(lambda: {"roles": set(), "times": 0, "money": 0.0, "diamonds": 0})
act_stat = defaultdict(lambda: {"roles": set(), "times": 0, "money": 0.0})
for r in item_role:
    s, act, gift = str(r["server_id"]), str(r["act_id"]), str(r["gift_id"])
    it = item_stat[(s, act, gift)]
    it["roles"].add(r["role_id"])
    it["times"] += int(r["times"])
    it["money"] += float(r["money"] or 0)
    it["diamonds"] += int(r["diamonds"] or 0)
    a = act_stat[(s, act)]
    a["roles"].add(r["role_id"])
    a["times"] += int(r["times"])
    a["money"] += float(r["money"] or 0)

# 等级分布
lvl_stat = {s: {"n": 0, "max": 0, "b": [0, 0, 0, 0, 0]} for s in ORDER}
for r in level_rows:
    s = str(r["server_id"])
    lvl = int(r["lvl"] or 0)
    st = lvl_stat[s]
    st["n"] += 1
    st["max"] = max(st["max"], lvl)
    bi = 0 if lvl < 20 else 1 if lvl < 40 else 2 if lvl < 60 else 3 if lvl < 80 else 4
    st["b"][bi] += 1

# 消耗聚合: (server, b_id) -> roles/times/diamonds
consume_stat = defaultdict(lambda: {"roles": set(), "times": 0, "diamonds": 0})
for r in consume_role:
    key = (str(r["server_id"]), int(r["b_id"]))
    cs = consume_stat[key]
    cs["roles"].add(r["role_id"])
    cs["times"] += int(r["times"])
    cs["diamonds"] += int(r["cost_diamonds"] or 0)

# ================= 写 CSV =================
write_csv(
    "01_新服概况.csv",
    ["批次", "server_id", "服名", "开服时间", "渠道opgame", "开服规律"],
    [[SERVER_META[s]["batch"], s, SERVER_META[s]["name"], SERVER_META[s]["open"], "2251", "每月6日一服"] for s in ORDER],
)

write_csv(
    "02_注册汇总.csv",
    ["批次", "服名", "server_id", "累计注册", "首7日注册"],
    [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s, reg_total[s], first7_reg[s]] for s in ORDER],
)

write_csv(
    "03_付费汇总.csv",
    ["批次", "服名", "server_id", "付费人数", "累计注册", "付费率%", "充值总额(美元)", "充值总钻石", "充值笔数", "ARPPU(美元)"],
    [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s,
      len(pay_stat[s]["roles"]), reg_total[s],
      round(len(pay_stat[s]["roles"]) / reg_total[s] * 100, 1) if reg_total[s] else 0,
      round(pay_stat[s]["money"], 2), pay_stat[s]["diamonds"], pay_stat[s]["times"],
      round(pay_stat[s]["money"] / len(pay_stat[s]["roles"]), 2) if pay_stat[s]["roles"] else 0] for s in ORDER],
)

write_csv(
    "04_首7日对齐对比.csv",
    ["批次", "服名", "server_id", "首7日注册", "首7日付费(美元)", "首7日ARPU(美元)"],
    [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s,
      first7_reg[s], round(first7_pay[s], 2),
      round(first7_pay[s] / first7_reg[s], 2) if first7_reg[s] else 0] for s in ORDER],
)

type_name = {"1": "购买钻石(普通充值)", "2": "购买商品(直购)"}
rows = []
for s in ORDER:
    for pt in ("1", "2"):
        t = type_stat.get((s, pt))
        if t:
            rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s,
                         type_name[pt], len(t["roles"]), t["times"], round(t["money"], 2)])
write_csv("05_付费类型拆分.csv", ["批次", "服名", "server_id", "付费类型", "付费人数", "笔数", "金额(美元)"], rows)

write_csv(
    "06_等级分布_近7天活跃.csv",
    ["批次", "服名", "server_id", "近7天活跃人数", "最高等级", "1-19级", "20-39级", "40-59级", "60-79级", "80级以上", "活跃/累计注册%"],
    [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s,
      lvl_stat[s]["n"], lvl_stat[s]["max"], *lvl_stat[s]["b"],
      round(lvl_stat[s]["n"] / reg_total[s] * 100, 1) if reg_total[s] else 0] for s in ORDER],
)

rows = []
for s in ORDER:
    tops = sorted(
        [(bid, cs) for (srv, bid), cs in consume_stat.items() if srv == s],
        key=lambda kv: -kv[1]["diamonds"],
    )[:15]
    for i, (bid, cs) in enumerate(tops, 1):
        rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s, i,
                     bid, glog_name(bid), len(cs["roles"]), cs["times"], cs["diamonds"]])
write_csv("07_钻石消耗去向TOP15.csv", ["批次", "服名", "server_id", "排名", "b_id", "来源(GLOG_SOURCE)", "消耗人数", "消耗次数", "消耗钻石"], rows)

rows = []
for s in ORDER:
    tops = sorted(
        [(act, a) for (srv, act), a in act_stat.items() if srv == s],
        key=lambda kv: -kv[1]["money"],
    )
    for i, (act, a) in enumerate(tops, 1):
        rows.append([SERVER_META[s]["batch"], SERVER_META[s]["name"], s, i,
                     act, DIRECT_ACT_NAME.get(act, ""), len(a["roles"]), a["times"], round(a["money"], 2)])
write_csv("08_直购活动分布_分服.csv", ["批次", "服名", "server_id", "排名", "活动ID", "活动名称", "付费人数", "购买笔数", "金额(美元)"], rows)

item_all = defaultdict(lambda: {"roles": set(), "times": 0, "money": 0.0, "diamonds": 0})
for (s, act, gift), it in item_stat.items():
    a = item_all[(act, gift)]
    a["roles"] |= it["roles"]  # 近似: 不同服 role_id 不重合
    a["times"] += it["times"]
    a["money"] += it["money"]
    a["diamonds"] += it["diamonds"]
rows = [[DIRECT_ACT_NAME.get(act, act), gift, len(a["roles"]), a["times"], round(a["money"], 2), a["diamonds"]]
        for (act, gift), a in sorted(item_all.items(), key=lambda kv: -kv[1]["money"])]
write_csv("09_直购商品TOP_全服合计.csv", ["直购活动", "商品ID(gift_id)", "付费人数", "购买笔数", "金额(美元)", "购得钻石"], rows)

rows = [[SERVER_META[s]["batch"], SERVER_META[s]["name"], s, DIRECT_ACT_NAME.get(act, act), gift,
         len(it["roles"]), it["times"], round(it["money"], 2), it["diamonds"]]
        for s in ORDER
        for (srv, act, gift), it in sorted(item_stat.items(), key=lambda kv: -kv[1]["money"]) if srv == s]
write_csv("10_直购商品TOP_分服.csv", ["批次", "服名", "server_id", "直购活动", "商品ID", "付费人数", "购买笔数", "金额(美元)", "购得钻石"], rows)

write_csv(
    "11_每日新增注册.csv",
    ["批次", "服名", "server_id", "日期(ds)", "新增注册"],
    [[SERVER_META[str(r["server_id"])]["batch"], SERVER_META[str(r["server_id"])]["name"], r["server_id"],
      r["ds"], r["new_roles"]] for r in daily_reg],
)

write_csv(
    "12_每日付费趋势.csv",
    ["批次", "服名", "server_id", "日期(ds)", "付费人数", "付费金额(美元)"],
    [[SERVER_META[str(r["server_id"])]["batch"], SERVER_META[str(r["server_id"])]["name"], r["server_id"],
      r["ds"], r["payers"], r["money"]] for r in daily_pay],
)

# ================= 合并 xlsx =================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SHEETS = [
    ("01_新服概况.csv", "新服概况"),
    ("02_注册汇总.csv", "注册汇总"),
    ("03_付费汇总.csv", "付费汇总"),
    ("04_首7日对齐对比.csv", "首7日对齐对比"),
    ("05_付费类型拆分.csv", "付费类型拆分"),
    ("06_等级分布_近7天活跃.csv", "等级分布(近7天活跃)"),
    ("07_钻石消耗去向TOP15.csv", "钻石消耗去向TOP15"),
    ("08_直购活动分布_分服.csv", "直购活动分布(分服)"),
    ("09_直购商品TOP_全服合计.csv", "直购商品TOP(全服合计)"),
    ("10_直购商品TOP_分服.csv", "直购商品TOP(分服)"),
    ("11_每日新增注册.csv", "每日新增注册"),
    ("12_每日付费趋势.csv", "每日付费趋势"),
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")

wb = Workbook()
wb.remove(wb.active)
for csv_name, sheet_name in SHEETS:
    with open(OUT_DIR / csv_name, newline="", encoding="utf-8-sig") as f:
        data = list(csv.reader(f))
    ws = wb.create_sheet(title=sheet_name)
    for r in data:
        converted = []
        for cell in r:
            for cast in (int, float):
                try:
                    converted.append(cast(cell))
                    break
                except (ValueError, TypeError):
                    continue
            else:
                converted.append(cell)
        ws.append(converted)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(ws.columns, 1):
        width = max((sum(2 if ord(ch) > 127 else 1 for ch in str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 40)

xlsx_path = OUT_DIR / "160新服分析_5月6月.xlsx"
wb.save(xlsx_path)
print(f"\n合并完成 -> {xlsx_path}")

# 只保留合并 xlsx, 删除中间 CSV
for csv_name, _ in SHEETS:
    (OUT_DIR / csv_name).unlink(missing_ok=True)

# ================= 摘要 =================
print("\n===== 摘要 =====")
for s in ORDER:
    m = SERVER_META[s]
    p = pay_stat[s]
    print(
        f"{m['batch']} {m['name']}: 注册={reg_total[s]} 首7日={first7_reg[s]} | "
        f"付费{len(p['roles'])}人 ${round(p['money'], 2)} | 首7日付费=${round(first7_pay[s], 2)} | "
        f"近7天活跃={lvl_stat[s]['n']} 最高级={lvl_stat[s]['max']}"
    )
for s in ORDER:
    m = SERVER_META[s]
    dp_roles = set()
    dp_money = 0.0
    for (srv, act), a in act_stat.items():
        if srv == s:
            dp_roles |= a["roles"]
            dp_money += a["money"]
    t1 = type_stat.get((s, "1"), {"money": 0.0})
    print(f"{m['batch']} {m['name']}: 普通充值=${round(t1['money'], 2)} | 直购 {len(dp_roles)}人 ${round(dp_money, 2)}")

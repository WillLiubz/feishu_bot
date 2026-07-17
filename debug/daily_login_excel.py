import json
import re
import sys
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent / "app"))

import config
import configdb

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
NAME_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def query(game_id, sql, database=None, max_rows=5000):
    gc = config.game_config(game_id)
    cdb = gc.config_db
    cfg = {
        "host": cdb["host"],
        "port": cdb["port"],
        "user": cdb["user"],
        "password": cdb["password"],
        "database": database or cdb["database"],
        "charset": cdb.get("charset", "utf8mb4"),
        "connect_timeout": cdb.get("connect_timeout", 5),
        "read_timeout": cdb.get("read_timeout", 30),
    }
    clean = configdb.sanitize(sql, max_rows=max_rows)
    return configdb.query(cfg, clean, max_rows=max_rows)


def fetch_item_names(item_ids):
    if not item_ids:
        return {}
    cfg = config.game_config(39).config_db
    db = cfg.get("static_database", cfg["database"])
    id_list = ",".join(str(i) for i in item_ids)
    rows = query(39, f"SELECT id, name FROM static_item WHERE id IN ({id_list})", database=db)
    return {r["id"]: r["name"] for r in rows}


def parse_reward_field(text):
    """Parse reward/cost string like 'items:990003275|2|0;items:56006|50|0'."""
    items = []
    if not text:
        return items
    for part in text.split(";"):
        part = part.strip()
        if not part:
            continue
        if part.startswith("items:"):
            rest = part[6:]
            pieces = rest.split("|")
            item_id = pieces[0]
            qty = pieces[1] if len(pieces) > 1 else ""
            bind = pieces[2] if len(pieces) > 2 else ""
            items.append({"type": "items", "item_id": item_id, "qty": qty, "bind": bind})
        elif part.startswith("cash:"):
            items.append({"type": "cash", "value": part[5:], "item_id": "", "qty": "", "bind": ""})
        else:
            items.append({"type": "other", "value": part, "item_id": "", "qty": "", "bind": ""})
    return items


def build_item_columns(max_items=5):
    cols = []
    for i in range(1, max_items + 1):
        cols.extend([
            f"item_id_{i}",
            f"item_name_{i}",
            f"item_qty_{i}",
            f"item_bind_{i}",
        ])
    return cols


def get_reward_row(act, reward, item_names, max_items=5):
    """Return dict for one reward row."""
    row = {
        "activity_id": act["id"],
        "activity_name": act["name"],
        "start_time": act["start_time"],
        "end_time": act["end_time"],
        "start_get_time": act["start_get_time"],
        "end_get_time": act["end_get_time"],
        "panel_stime": act["panel_stime"],
        "panel_etime": act["panel_etime"],
        "status": act["status"],
        "reward_id": reward["id"],
        "reward_type": reward["reward_type"],
        "type_value": reward["type_value"],
        "type_times": reward["type_times"],
        "button_type": reward["button_type"],
        "cond_desc": reward["cond_desc"],
        "func_type": reward["func_type"],
        "func_content": reward["func_content"],
        "relation": reward["relation"],
        "send_msg": reward["send_msg"],
        "cost": reward["cost"],
        "reward": reward["reward"],
    }
    parsed = parse_reward_field(reward.get("reward", ""))
    for i in range(1, max_items + 1):
        idx = i - 1
        if idx < len(parsed):
            p = parsed[idx]
            row[f"item_id_{i}"] = p.get("item_id", "")
            row[f"item_name_{i}"] = item_names.get(int(p["item_id"]), "") if p.get("item_id") and p["item_id"].isdigit() else ""
            row[f"item_qty_{i}"] = p.get("qty", "")
            row[f"item_bind_{i}"] = p.get("bind", "")
        else:
            row[f"item_id_{i}"] = ""
            row[f"item_name_{i}"] = ""
            row[f"item_qty_{i}"] = ""
            row[f"item_bind_{i}"] = ""
    return row


def create_activity_compare_sheet(ws, may_acts, jun_acts):
    """Create side-by-side activity comparison sheet."""
    activity_cols = [
        "activity_id", "name", "start_time", "end_time", "start_get_time", "end_get_time",
        "panel_stime", "panel_etime", "status", "reward_ids",
    ]
    # 表头：5月列 + 6月列
    headers = []
    for col in activity_cols:
        headers.append(f"5月_{col}")
    for col in activity_cols:
        headers.append(f"6月_{col}")

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).font = Font(bold=True)

    # 对齐活动：按开始日期顺序
    may_sorted = sorted(may_acts, key=lambda x: x["start_time"])
    jun_sorted = sorted(jun_acts, key=lambda x: x["start_time"])

    for i in range(max(len(may_sorted), len(jun_sorted))):
        may_act = may_sorted[i] if i < len(may_sorted) else {}
        jun_act = jun_sorted[i] if i < len(jun_sorted) else {}
        row = []
        for col in activity_cols:
            row.append(may_act.get(col, ""))
        for col in activity_cols:
            row.append(jun_act.get(col, ""))
        ws.append(row)
        r = ws.max_row
        # 标黄不一致
        for idx, col in enumerate(activity_cols):
            may_val = may_act.get(col, "")
            jun_val = jun_act.get(col, "")
            if may_val != jun_val:
                ws.cell(row=r, column=idx + 1).fill = YELLOW_FILL
                ws.cell(row=r, column=idx + 1 + len(activity_cols)).fill = YELLOW_FILL

    # 调整列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22


def create_reward_compare_sheet(ws, may_acts, jun_acts, item_names, max_items=5):
    """Create side-by-side reward comparison sheet."""
    base_reward_cols = [
        "activity_id", "activity_name", "start_time", "end_time",
        "reward_id", "reward_type", "type_value", "type_times",
        "button_type", "cond_desc", "func_type", "func_content",
        "relation", "send_msg", "cost", "reward",
    ]
    item_cols = build_item_columns(max_items)
    reward_cols = base_reward_cols + item_cols

    headers = []
    for col in reward_cols:
        headers.append(f"5月_{col}")
    for col in reward_cols:
        headers.append(f"6月_{col}")

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 第二行：道具名称提示（仅在 item_name 列）
    name_row = [""] * len(headers)
    for side_idx, prefix in [(0, "5月"), (len(reward_cols), "6月")]:
        for i in range(1, max_items + 1):
            col_idx = side_idx + reward_cols.index(f"item_name_{i}") + 1
            name_row[col_idx - 1] = "道具名称"
    ws.append(name_row)
    for c in range(1, len(headers) + 1):
        if name_row[c - 1] == "道具名称":
            ws.cell(row=2, column=c).fill = NAME_FILL
            ws.cell(row=2, column=c).font = Font(italic=True, color="006100")

    # 对齐 reward：按活动顺序，每个活动内的 reward 按 id 顺序
    may_sorted = sorted(may_acts, key=lambda x: x["start_time"])
    jun_sorted = sorted(jun_acts, key=lambda x: x["start_time"])

    may_rewards = []
    for act in may_sorted:
        for r in sorted(act.get("rewards", []), key=lambda x: x["id"]):
            may_rewards.append((act, r))
    jun_rewards = []
    for act in jun_sorted:
        for r in sorted(act.get("rewards", []), key=lambda x: x["id"]):
            jun_rewards.append((act, r))

    for i in range(max(len(may_rewards), len(jun_rewards))):
        may_pair = may_rewards[i] if i < len(may_rewards) else (None, None)
        jun_pair = jun_rewards[i] if i < len(jun_rewards) else (None, None)
        may_row = get_reward_row(may_pair[0], may_pair[1], item_names, max_items) if may_pair[0] else {c: "" for c in reward_cols}
        jun_row = get_reward_row(jun_pair[0], jun_pair[1], item_names, max_items) if jun_pair[0] else {c: "" for c in reward_cols}

        row = [may_row[c] for c in reward_cols] + [jun_row[c] for c in reward_cols]
        ws.append(row)
        r = ws.max_row

        # 标黄不一致（跳过 item_name 列，因为它只是展示）
        for idx, col in enumerate(reward_cols):
            if col.startswith("item_name_"):
                continue
            may_val = may_row[col]
            jun_val = jun_row[col]
            if may_val != jun_val:
                ws.cell(row=r, column=idx + 1).fill = YELLOW_FILL
                ws.cell(row=r, column=idx + 1 + len(reward_cols)).fill = YELLOW_FILL

    # 调整列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18


def main():
    data = json.loads(Path("debug/output/activity_39_may_jun_2026_raw.json").read_text(encoding="utf-8"))

    may_acts = [a for a in data["may"] if a["name"] == "每日登陆"]
    jun_acts = [a for a in data["june"] if a["name"] == "每日登陆"]

    # 收集所有道具ID
    item_ids = set()
    for acts in [may_acts, jun_acts]:
        for a in acts:
            for r in a.get("rewards", []):
                for p in parse_reward_field(r.get("reward", "")):
                    if p.get("item_id") and p["item_id"].isdigit():
                        item_ids.add(int(p["item_id"]))

    item_names = fetch_item_names(item_ids)

    wb = Workbook()

    # 删除默认sheet
    wb.remove(wb.active)

    ws_act = wb.create_sheet("活动对比")
    create_activity_compare_sheet(ws_act, may_acts, jun_acts)

    ws_rew = wb.create_sheet("奖励对比")
    create_reward_compare_sheet(ws_rew, may_acts, jun_acts, item_names, max_items=5)

    out_path = Path("debug/output/activity_39_daily_login_may_jun_2026.xlsx")
    wb.save(out_path)
    print(f"Excel saved to {out_path}")


if __name__ == "__main__":
    main()

"""把 debug/output/39_new_servers/ 下的 11 个 CSV 合并为一个多 sheet Excel。

输出: debug/output/39_new_servers/39新服分析_5月6月.xlsx
"""
import csv
import io
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SRC_DIR = Path(__file__).parent / "output" / "39_new_servers"
OUT_FILE = SRC_DIR / "39新服分析_5月6月.xlsx"

# (csv 文件名, sheet 名) — sheet 名不能超过 31 字符
SHEETS = [
    ("01_新服概况.csv", "新服概况"),
    ("02_注册汇总.csv", "注册汇总"),
    ("03_付费汇总.csv", "付费汇总"),
    ("04_首7日对齐对比.csv", "首7日对齐对比"),
    ("05_付费类型拆分.csv", "付费类型拆分"),
    ("06_等级分布_近7天活跃.csv", "等级分布(近7天活跃)"),
    ("07_钻石消费系统TOP15.csv", "钻石消费系统TOP15"),
    ("08_直购礼包TOP_全服合计.csv", "直购礼包TOP(全服合计)"),
    ("09_直购礼包TOP_分服.csv", "直购礼包TOP(分服)"),
    ("10_每日新增注册.csv", "每日新增注册"),
    ("11_每日付费趋势.csv", "每日付费趋势"),
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")

wb = Workbook()
wb.remove(wb.active)

for csv_name, sheet_name in SHEETS:
    path = SRC_DIR / csv_name
    if not path.exists():
        print(f"[skip] {csv_name} 不存在")
        continue
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        continue
    ws = wb.create_sheet(title=sheet_name)
    for r in rows:
        # 尽量把数字转回数值类型,方便 Excel 里求和/排序
        converted = []
        for cell in r:
            try:
                converted.append(int(cell))
                continue
            except (ValueError, TypeError):
                pass
            try:
                converted.append(float(cell))
                continue
            except (ValueError, TypeError):
                pass
            converted.append(cell)
        ws.append(converted)
    # 表头样式 + 冻结首行 + 自适应列宽
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(ws.columns, 1):
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        # 中文按 2 个字符宽度估算
        width = max((sum(2 if ord(ch) > 127 else 1 for ch in str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 40)
    print(f"[OK] sheet [{sheet_name}]  {len(rows) - 1} 行")

wb.save(OUT_FILE)
print(f"\n合并完成 -> {OUT_FILE}")
